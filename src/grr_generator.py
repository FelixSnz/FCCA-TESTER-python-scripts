
import os
import glob
import pandas as pd
from openpyxl import load_workbook


from loghandler import Log
import logtools
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.worksheet import Worksheet
import configparser
from datetime import datetime

from utils import range_alpha

config = configparser.ConfigParser()
config.read('config\settings.ini')

GRRsSavePath = config["Paths"]["GRRs"]

log_master_path = config["Paths"]['LogMaster']

logs_path = r"resources\OUT"

redFill = PatternFill(start_color='EE1111',
                end_color='EE1111',
                fill_type='solid')

blankFill = PatternFill(start_color='FFFFFF',
                    end_color='FFFFFF',
                    fill_type='solid')

greenFill = PatternFill(start_color='11EE11',
                    end_color='11EE11',
                    fill_type='solid')

orangeFill = PatternFill(start_color='FFA500',
                    end_color='FFA500',
                    fill_type='solid')


def main():

    grr_data = []
    themoking_grr = GRR("..\\test.xlsx")

    for log_fname in glob.iglob(logs_path + '/**', recursive=True):
        if os.path.isfile(log_fname):
            themoking_grr.append_log(log_fname)
            print(log_fname)

class GRR:
    def __init__(self, save_name, file_exists=False) -> None:
        self.log_master = Log(log_master_path)
        self.path =  save_name
        self.tests_info = ["SerialNumber", "Date", "Time", "Status", " "] + logtools.get_all_test_names(self.log_master)
        self.low_limits =  ["", "", "", "", "LOW LIMITS"] + logtools.get_all_limits(self.log_master, "low")
        self.high_limits = ["", "", "", "", "HIGH LIMITS"] + logtools.get_all_limits(self.log_master, "high")
        self.column_amount = len(self.tests_info)
        if not file_exists:
            self.create_blank(save_name)
    
    def get_cell_limits(self, row_amount):
        limits = (self.low_limits[5:], self.high_limits[5:])
        cells_limits = {}
        alphas = range_alpha("G", self.column_amount - 5)
        for idx, alpha in enumerate(alphas):
            key = f"{alpha}{4}:{alpha}{row_amount}"
            val = [limits[0][idx], limits[1][idx]]
            cells_limits[key] = val
        return cells_limits
    

    def get_log_data(self, log_path):
        log_data = []
        if  log_path != None and os.path.isfile(log_path):
            log = Log(log_path)
            log_data = [log.serial, log.date, log.time, log.get_status(), " "]
            with open(log_path) as log_file:
                lines = log_file.readlines()
                passed_lines = []
                for line in lines:
                    passed_lines.append(line)
                    result_info = log.get_result_info(line)
                    if result_info != None:
                        line_status = result_info["status"]
                        if result_info["type"] == 'numeric':
                            result = result_info["measurement"]
                        elif result_info["type"] == 'binary':
                            result = line_status
                        log_data.append(result)
                self.fill_data(log_data)
        return log_data


    def append_log(self, log_path):
        log_data = self.get_log_data(log_path)
        
        writer = pd.ExcelWriter(self.path, engine='openpyxl', mode="a", if_sheet_exists='overlay')
        writer.book = load_workbook(self.path)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        reader = pd.read_excel(self.path)
        grr_data_df = pd.DataFrame(data=[log_data], columns=[self.tests_info, self.low_limits, self.high_limits], index=[writer.book.worksheets[0].max_row-2])
        
        grr_data_df.to_excel(writer,index=True,header=False,startrow=len(reader))
        writer.close()
        self.apply_cell_format()
    
    def apply_status_color(self, ws:Worksheet):
        row_amount = ws.max_row
        for n in range(4, row_amount+ 1):
            ws.conditional_formatting.add(f"E{n}",FormulaRule(formula=[f'E{n}="PASS"'], stopIfTrue=False, fill=greenFill))
            ws.conditional_formatting.add(f"E{n}",FormulaRule(formula=[f'E{n}="INCOMPLETE"'], stopIfTrue=False, fill=orangeFill))
            ws.conditional_formatting.add(f"E{n}",FormulaRule(formula=[f'E{n}="FAIL"'], stopIfTrue=False, fill=redFill))


    def apply_result_color(self, ws:Worksheet):
        row_amount = ws.max_row
        cells_lims = self.get_cell_limits(row_amount)
        for cellsref, range_ in cells_lims.items():
            print(cellsref)
            ws.conditional_formatting.add(cellsref, FormulaRule(formula=[f'ISBLANK({cellsref.split(":")[0]})'], stopIfTrue=False, fill=blankFill))
            ws.conditional_formatting.add(cellsref, CellIsRule(operator='notBetween', formula=range_, stopIfTrue=False, fill=redFill))
            ws.conditional_formatting.add(cellsref, FormulaRule(formula=[f'NOT(ISERROR(SEARCH("FAIL",{cellsref})))'], stopIfTrue=False, fill=redFill))
    
    def apply_cell_format(self):
        work_book = load_workbook(self.path)
        worksheet = work_book.worksheets[0]
        self.apply_status_color(worksheet)
        self.apply_result_color(worksheet)
        work_book.save(self.path)

                
    def create_blank(self, xlsx_save_name):
        writer = pd.ExcelWriter(xlsx_save_name, engine='xlsxwriter')
        writer.save()
        df = pd.DataFrame(data=None, columns=[self.tests_info, self.low_limits, self.high_limits])
        writer = pd.ExcelWriter(xlsx_save_name, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Sheet1', index=True)
        writer.save()
        print(xlsx_save_name)

    
    def fill_data(self, data):
        if len(data) < self.column_amount:
            diff = self.column_amount - len(data)
            data += [None] * diff


def generate_grr_savename():
    return GRRsSavePath + "\\GRR_" + str(datetime.today()).split(" ")[0] + ".xlsx"







if __name__ == "__main__":
    main()
